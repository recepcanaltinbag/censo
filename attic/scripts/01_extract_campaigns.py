#!/usr/bin/env python3
"""
Extract the four Ergene sampling campaigns into tidy CSV.

Inputs  (read-only):  Data/Sampling/Sampling/*_Sampling-*.xlsx
Outputs (generated):  derived/processed/stations.csv
                      derived/processed/measurements.csv
                      derived/processed/parameters.csv
                      derived/interim/extraction_report.txt

Design notes
------------
* Censoring is preserved, never substituted. A cell reading "<0.05" becomes
  value_num=0.05 with qualifier="<". Downstream code decides what to do with it;
  this script never invents a number (cf. Helsel 2006).
* Every path is relative to the repository root, derived from __file__.
* Deterministic: inputs are sorted, no dict-order dependence, no randomness.

Usage:  python scripts/01_extract_campaigns.py
"""

from __future__ import annotations

import csv
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "Data" / "Sampling" / "Sampling"
OUT = ROOT / "derived" / "processed"
INTERIM = ROOT / "derived" / "interim"

# Campaign file -> (campaign_id, season, nominal year-month).
#
# All four filenames carry a 2018 EXPORT date, which is not the sampling date.
# The chronological order is fixed by the in-sheet `Tarih` row and confirmed
# independently by the survey's own publication (Emadian et al. 2021, STOTEN
# 758:143656), which gives: Summer 2017 (8-11 Aug), Fall 2017 (14-16 Nov),
# Winter 2018 (12-16 Feb), Spring 2018 (14-15 May).
#
# An earlier version of this file numbered November as C1 and August as C4
# (2018), putting the campaigns out of order and mislabelling the August year.
# Seasonal ordering matters for the persistence analysis, so it is fixed here.
CAMPAIGNS = {
    "August_Sampling-2018_08_01.xlsx": ("C1", "summer", "2017-08"),
    "November_Sampling-2018_08_01.xlsx": ("C2", "autumn", "2017-11"),
    "February_Sampling-2018_08_01.xlsx": ("C3", "winter", "2018-02"),
    "May_Sampling-2018_08_08.xlsx": ("C4", "spring", "2018-05"),
}

SHEET = "Analizler"

# Row labels (column A) that carry station metadata rather than measurements.
META_ROWS = {"Enlem", "Boylam", "Yükseklik", "Tarih", "Zaman"}

# Section headers in column A. They have no unit and no values; they classify the
# rows that follow, which is exactly the analyte grouping we need downstream.
SECTIONS = {
    "Konvansiyonel": "conventional",
    "Toksisite": "toxicity",
    "Debi": "flow",
    "Metals": "metal",
    "Mikrokirleticiler": "micropollutant",
}

# Rows that are spreadsheet bookkeeping, not measurements.
SKIP_PARAMS = {"COUNT", "COUNT>ÇKS", "Parametre"}

DMS_RE = re.compile(
    r"""^\s*(?P<deg>\d+(?:[.,]\d+)?)\s*[°ºd]\s*
         (?:(?P<min>\d+(?:[.,]\d+)?)\s*['′m]\s*)?
         (?:(?P<sec>\d+(?:[.,]\d+)?)\s*(?:["″”]|'')\s*)?
         (?P<hemi>[NSEWnsew])?\s*$""",
    re.VERBOSE,
)

# Leading qualifier on a reported concentration.
QUAL_RE = re.compile(r"^\s*(?P<q><=|>=|<|>|≤|≥)\s*(?P<rest>.+)$")

# Textual non-detect markers seen in environmental spreadsheets.
NON_DETECT = {
    "nd", "n.d.", "n.d", "bdl", "<lod", "lod", "yok", "tespit edilemedi", "-", "—",
}


def norm(s) -> str:
    """Normalise a cell label: NFC, collapse whitespace, strip."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(s))).strip()


def parse_dms(raw) -> float | None:
    """Convert a DMS string such as 41°47'30.7\" to decimal degrees.

    Returns None if unparseable. Does NOT range-check; validity is reported
    separately so that bad records surface instead of vanishing.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = norm(raw).replace(",", ".")
    m = DMS_RE.match(text)
    if not m:
        try:
            return float(text)
        except ValueError:
            return None
    deg = float(m.group("deg"))
    minutes = float(m.group("min") or 0.0)
    seconds = float(m.group("sec") or 0.0)
    value = deg + minutes / 60.0 + seconds / 3600.0
    if (m.group("hemi") or "").upper() in {"S", "W"}:
        value = -value
    return value


def dms_is_valid(raw) -> bool:
    """True when the minute and second fields are within range."""
    if raw is None or isinstance(raw, (int, float)):
        return True
    m = DMS_RE.match(norm(raw).replace(",", "."))
    if not m:
        return False
    minutes = float(m.group("min") or 0.0)
    seconds = float(m.group("sec") or 0.0)
    return minutes < 60.0 and seconds < 60.0


def parse_value(raw):
    """Parse a measurement cell.

    Returns (value_num, qualifier, raw_text) where qualifier is one of
    '<', '>', '<=', '>=', 'nd' or '' (plain number).
    value_num is the *reported bound*, never a substituted estimate.
    """
    if raw is None:
        return None, "", ""
    if isinstance(raw, (int, float)):
        return float(raw), "", str(raw)

    text = norm(raw)
    if not text:
        return None, "", ""
    if text.lower() in NON_DETECT:
        return None, "nd", text

    m = QUAL_RE.match(text)
    qual = ""
    body = text
    if m:
        qual = m.group("q").replace("≤", "<=").replace("≥", ">=")
        body = m.group("rest")

    body = body.replace(",", ".").replace(" ", "")
    try:
        return float(body), qual, text
    except ValueError:
        return None, "unparsed", text


def is_station_column(label) -> bool:
    """Station columns are numbered 1..84; trailing columns are aggregates."""
    text = norm(label)
    return bool(text) and text.replace(".0", "").isdigit()


def extract_campaign(path: Path, campaign_id: str, season: str, nominal: str,
                     stations: dict, measurements: list, parameters: dict,
                     report: list) -> None:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        report.append(f"  !! sheet '{SHEET}' missing in {path.name}")
        wb.close()
        return

    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    # column index -> station id
    cols = {i: norm(v).replace(".0", "")
            for i, v in enumerate(header)
            if i >= 2 and is_station_column(v)}
    report.append(f"  stations in header: {len(cols)}")

    meta = {}
    current_group = "unassigned"
    n_meas = 0
    coord_problems = []

    for row in rows[1:]:
        param = norm(row[0] if row else None)
        if not param or param in SKIP_PARAMS:
            continue

        unit = norm(row[1]) if len(row) > 1 else ""
        values_present = sum(1 for i in cols if i < len(row) and row[i] is not None)

        # Section header: no unit, no values.
        if param in SECTIONS and not unit and values_present == 0:
            current_group = SECTIONS[param]
            continue

        # Station metadata rows.
        if param in META_ROWS:
            meta[param] = {sid: row[i] if i < len(row) else None
                           for i, sid in cols.items()}
            continue

        # A measurement row.
        parameters.setdefault(param, {"parameter": param, "unit": unit,
                                      "group": current_group, "campaigns": set()})
        parameters[param]["campaigns"].add(campaign_id)
        if not parameters[param]["unit"] and unit:
            parameters[param]["unit"] = unit

        for i, sid in cols.items():
            raw = row[i] if i < len(row) else None
            if raw is None:
                continue
            value, qual, raw_text = parse_value(raw)
            if value is None and qual == "":
                continue
            measurements.append({
                "campaign": campaign_id,
                "station": sid,
                "parameter": param,
                "group": current_group,
                "unit": unit,
                "value_num": "" if value is None else repr(value),
                "qualifier": qual,
                "raw": raw_text,
            })
            n_meas += 1

    # Fold station metadata into the global table.
    for sid in cols.values():
        lat_raw = meta.get("Enlem", {}).get(sid)
        lon_raw = meta.get("Boylam", {}).get(sid)
        if not dms_is_valid(lat_raw):
            coord_problems.append(f"station {sid}: invalid latitude {lat_raw!r}")
        if not dms_is_valid(lon_raw):
            coord_problems.append(f"station {sid}: invalid longitude {lon_raw!r}")

        key = (campaign_id, sid)
        stations[key] = {
            "campaign": campaign_id,
            "season": season,
            "nominal_period": nominal,
            "station": sid,
            "lat_dms": norm(lat_raw),
            "lon_dms": norm(lon_raw),
            "lat": "" if (v := parse_dms(lat_raw)) is None else f"{v:.6f}",
            "lon": "" if (v := parse_dms(lon_raw)) is None else f"{v:.6f}",
            "coord_valid": str(dms_is_valid(lat_raw) and dms_is_valid(lon_raw)),
            "altitude_m": norm(meta.get("Yükseklik", {}).get(sid)),
            "date": norm(meta.get("Tarih", {}).get(sid)),
            "time": norm(meta.get("Zaman", {}).get(sid)),
        }

    report.append(f"  measurements: {n_meas}")
    if coord_problems:
        report.append(f"  COORDINATE PROBLEMS ({len(coord_problems)}):")
        report.extend(f"    - {p}" for p in coord_problems)


def main() -> int:
    if not SRC.is_dir():
        sys.exit(f"input directory not found: {SRC}")
    OUT.mkdir(parents=True, exist_ok=True)
    INTERIM.mkdir(parents=True, exist_ok=True)

    stations: dict = {}
    measurements: list = []
    parameters: dict = {}
    report: list = ["Ergene campaign extraction report", "=" * 40, ""]

    for fname in sorted(CAMPAIGNS):
        path = SRC / fname
        cid, season, nominal = CAMPAIGNS[fname]
        report.append(f"[{cid}] {fname}")
        if not path.exists():
            report.append("  !! FILE MISSING")
            continue
        extract_campaign(path, cid, season, nominal,
                         stations, measurements, parameters, report)
        report.append("")

    # ---- write outputs (sorted for determinism) ----
    with (OUT / "stations.csv").open("w", newline="", encoding="utf-8") as fh:
        fields = ["campaign", "season", "nominal_period", "station", "lat", "lon",
                  "lat_dms", "lon_dms", "coord_valid", "altitude_m", "date", "time"]
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for key in sorted(stations, key=lambda k: (k[0], int(k[1]))):
            w.writerow(stations[key])

    with (OUT / "measurements.csv").open("w", newline="", encoding="utf-8") as fh:
        fields = ["campaign", "station", "parameter", "group", "unit",
                  "value_num", "qualifier", "raw"]
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for m in sorted(measurements,
                        key=lambda r: (r["campaign"], int(r["station"]), r["parameter"])):
            w.writerow(m)

    with (OUT / "parameters.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["parameter", "unit", "group", "n_campaigns"])
        w.writeheader()
        for name in sorted(parameters):
            p = parameters[name]
            w.writerow({"parameter": p["parameter"], "unit": p["unit"],
                        "group": p["group"], "n_campaigns": len(p["campaigns"])})

    # ---- summary ----
    by_group: dict = {}
    by_qual: dict = {}
    for m in measurements:
        by_group[m["group"]] = by_group.get(m["group"], 0) + 1
        by_qual[m["qualifier"] or "plain"] = by_qual.get(m["qualifier"] or "plain", 0) + 1

    report += ["", "SUMMARY", "-" * 40,
               f"campaigns          : {len(CAMPAIGNS)}",
               f"station records    : {len(stations)}",
               f"distinct stations  : {len({k[1] for k in stations})}",
               f"parameters         : {len(parameters)}",
               f"measurements       : {len(measurements)}",
               "", "measurements by group:"]
    report += [f"  {g:18s} {n:7d}" for g, n in sorted(by_group.items(),
                                                      key=lambda x: -x[1])]
    report += ["", "measurements by qualifier:"]
    report += [f"  {q:18s} {n:7d}" for q, n in sorted(by_qual.items(),
                                                     key=lambda x: -x[1])]
    report += ["", "parameters by group:"]
    pg: dict = {}
    for p in parameters.values():
        pg[p["group"]] = pg.get(p["group"], 0) + 1
    report += [f"  {g:18s} {n:5d}" for g, n in sorted(pg.items(), key=lambda x: -x[1])]

    text = "\n".join(report)
    (INTERIM / "extraction_report.txt").write_text(text, encoding="utf-8")
    print(text)
    print(f"\nwrote: {OUT/'stations.csv'}, {OUT/'measurements.csv'}, "
          f"{OUT/'parameters.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

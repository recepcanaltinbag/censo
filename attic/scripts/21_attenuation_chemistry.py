#!/usr/bin/env python3
"""
Does the mass-balance classification agree with the chemistry?

WHY THIS IS A FALSIFICATION TEST, NOT A DECORATION
--------------------------------------------------
scripts/20_mass_balance.py splits each disappearance into two classes:

  masked_by_dilution    the upstream mass could still be present, merely spread
                        below the detection limit;
  attenuation_required  conservation is violated unless mass was genuinely lost.

Nothing in that split uses chemistry -- it is pure hydrology and arithmetic. So
chemistry supplies an independent check. Substances that really are removed from
river water are removed by known mechanisms: sorption to sediment and organic
matter, which rises steeply with hydrophobicity, and biodegradation. If the
split is detecting real behaviour, the `attenuation_required` substances should
be systematically MORE hydrophobic and LESS soluble than the masked ones.

If they are not, the split may be noise, and we would have to say so. The test
is run and reported either way; a null result is recorded, not discarded.

AVOIDING PSEUDO-REPLICATION
---------------------------
One substance contributes many reach x campaign comparisons, and they are not
independent -- they share the substance's own chemistry. Testing over
comparisons would inflate n several-fold and manufacture significance. The unit
here is therefore the SUBSTANCE: each gets the fraction of its decidable
comparisons that required attenuation, and that fraction is correlated against
its properties.

STATISTICS
----------
Spearman's rho (rank-based; the property distributions are skewed and the
sample is small) with an exact-ish permutation p-value, so no SciPy dependency
is introduced and the test can be re-run offline.

Inputs  : derived/processed/mass_balance.csv
          Data/Properties_report order_2018-07-04.xlsx
Outputs : derived/processed/attenuation_chemistry.csv
          eval/attenuation_chemistry.md

Usage:  python scripts/21_attenuation_chemistry.py [--iters 20000]
"""

from __future__ import annotations

import argparse
import csv
import math
import random
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "Data"
PROC = ROOT / "derived" / "processed"
EVAL = ROOT / "eval"
PROPS = DATA / "Properties_report order_2018-07-04.xlsx"

SEED = 20260803          # fixed: the permutation test must be reproducible


def fold(s: str) -> str:
    s = str(s)
    for a, b in (("ı", "i"), ("İ", "i"), ("ş", "s"), ("Ş", "s"), ("ğ", "g"),
                 ("Ğ", "g"), ("ç", "c"), ("Ç", "c"), ("ö", "o"), ("Ö", "o"),
                 ("ü", "u"), ("Ü", "u")):
        s = s.replace(a, b)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", s.lower())


def num(x):
    try:
        v = float(str(x).strip().replace(",", "."))
        return v if math.isfinite(v) else None
    except (TypeError, ValueError):
        return None


def ranks(xs):
    """Average ranks, so ties do not bias rho."""
    order = sorted(range(len(xs)), key=lambda i: xs[i])
    r = [0.0] * len(xs)
    i = 0
    while i < len(order):
        j = i
        while j + 1 < len(order) and xs[order[j + 1]] == xs[order[i]]:
            j += 1
        avg = (i + j) / 2.0 + 1.0
        for k in range(i, j + 1):
            r[order[k]] = avg
        i = j + 1
    return r


def spearman(x, y):
    rx, ry = ranks(x), ranks(y)
    n = len(x)
    mx, my = sum(rx) / n, sum(ry) / n
    num_ = sum((a - mx) * (b - my) for a, b in zip(rx, ry))
    den = math.sqrt(sum((a - mx) ** 2 for a in rx)
                    * sum((b - my) ** 2 for b in ry))
    return num_ / den if den else 0.0


def perm_p(x, y, iters, rng):
    """Two-sided permutation p-value for Spearman's rho."""
    obs = spearman(x, y)
    yy = list(y)
    hits = 0
    for _ in range(iters):
        rng.shuffle(yy)
        if abs(spearman(x, yy)) >= abs(obs) - 1e-12:
            hits += 1
    return obs, (hits + 1) / (iters + 1)


def read_properties():
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required:  pip install openpyxl")
    if not PROPS.exists():
        sys.exit(f"missing {PROPS}")
    wb = openpyxl.load_workbook(PROPS, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c or "").replace("\n", " ").strip() for c in rows[0]]

    def col(*needles):
        for i, h in enumerate(hdr):
            hl = h.lower()
            if all(n in hl for n in needles):
                return i
        return None

    i_name = col("chemical", "name")
    i_logp = col("logp")
    i_sol = col("solubility", "water")
    i_mw = col("chemical", "mw")
    out = {}
    for r in rows[1:]:
        if not r or i_name is None or r[i_name] is None:
            continue
        out[fold(r[i_name])] = {
            "name": str(r[i_name]).strip(),
            "logp": num(r[i_logp]) if i_logp is not None else None,
            "solubility_ppm": num(r[i_sol]) if i_sol is not None else None,
            "mw": num(r[i_mw]) if i_mw is not None else None,
        }
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--iters", type=int, default=20000)
    args = ap.parse_args()
    EVAL.mkdir(parents=True, exist_ok=True)
    rng = random.Random(SEED)

    mb_path = PROC / "mass_balance.csv"
    if not mb_path.exists():
        sys.exit("missing mass_balance.csv; run scripts/20_mass_balance.py first")
    with mb_path.open(encoding="utf-8") as fh:
        mb = [r for r in csv.DictReader(fh)
              if r["verdict"] in ("masked_by_dilution", "attenuation_required")]
    if not mb:
        sys.exit("no decidable comparisons in mass_balance.csv")

    props = read_properties()

    per_sub = defaultdict(lambda: {"n": 0, "att": 0})
    for r in mb:
        d = per_sub[r["analyte"]]
        d["n"] += 1
        d["att"] += 1 if r["verdict"] == "attenuation_required" else 0

    rows, unmatched = [], []
    for name, d in sorted(per_sub.items()):
        p = props.get(fold(name))
        if p is None:
            unmatched.append(name)
            continue
        rows.append({
            "analyte": name,
            "n_comparisons": d["n"],
            "n_attenuation_required": d["att"],
            "attenuation_fraction": f"{d['att']/d['n']:.4f}",
            "logp": "" if p["logp"] is None else f"{p['logp']:.3f}",
            "solubility_ppm": "" if p["solubility_ppm"] is None
                              else f"{p['solubility_ppm']:.6g}",
            "mw": "" if p["mw"] is None else f"{p['mw']:.2f}",
        })

    if rows:
        with (PROC / "attenuation_chemistry.csv").open(
                "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)

    tests = []
    for key, label, expect in (
            ("logp", "log P (hydrophobicity)",
             "positive: hydrophobic substances sorb and are removed"),
            ("solubility_ppm", "water solubility (ppm)",
             "negative: soluble substances stay in the water column"),
            ("mw", "molecular weight",
             "no strong prior")):
        xs, ys, names = [], [], []
        for r in rows:
            v = num(r[key])
            if v is None:
                continue
            if key == "solubility_ppm":
                if v <= 0:
                    continue          # log scale; zeros are 'insoluble' flags
                v = math.log10(v)
            xs.append(v)
            ys.append(float(r["attenuation_fraction"]))
            names.append(r["analyte"])
        if len(xs) < 6:
            tests.append((label, len(xs), None, None, expect,
                          "too few substances to test"))
            continue
        rho, p = perm_p(xs, ys, args.iters, rng)
        tests.append((label, len(xs), rho, p, expect, ""))

    L = []
    A = L.append
    A("# Does the chemistry agree with the mass balance?\n")
    A("Generated by `scripts/21_attenuation_chemistry.py`.\n")
    A("The split in `eval/mass_balance.md` uses only hydrology and arithmetic: "
      "no chemical property enters it. Chemistry is therefore an independent "
      "check. Substances genuinely removed from river water are removed by "
      "sorption, which rises with hydrophobicity, and by biodegradation. If "
      "the split detects real behaviour, the substances needing attenuation "
      "should be more hydrophobic and less soluble.\n")
    A("The unit of analysis is the **substance**, not the comparison: one "
      "substance contributes many reach-campaign comparisons that share its "
      "chemistry, and testing over those would inflate the sample and "
      "manufacture significance.\n")
    A(f"- substances with at least one decidable comparison: "
      f"**{len(per_sub)}**")
    A(f"- matched to a property record: **{len(rows)}**")
    if unmatched:
        A(f"- unmatched, excluded: {len(unmatched)} "
          f"({', '.join(sorted(unmatched)[:6])}"
          f"{' …' if len(unmatched) > 6 else ''})")
    A(f"- permutation iterations: {args.iters:,}, seed {SEED}\n")

    A("## Correlation with the fraction of comparisons requiring attenuation\n")
    A("| property | n | Spearman rho | p (permutation) | expected sign |")
    A("|---|---|---|---|---|")
    for label, n, rho, p, expect, note in tests:
        if rho is None:
            A(f"| {label} | {n} | — | — | {expect} ({note}) |")
        else:
            A(f"| {label} | {n} | **{rho:+.3f}** | {p:.4f} | {expect} |")
    A("")

    sig = [t for t in tests if t[2] is not None and t[3] is not None
           and t[3] < 0.05]
    logp = next((t for t in tests if t[0].startswith("log P")), None)
    if logp and logp[2] is not None:
        rho, p = logp[2], logp[3]
        if p < 0.05 and rho > 0:
            A(f"> **The chemistry agrees.** Hydrophobicity predicts which "
              f"disappearances mass balance says are real "
              f"(rho = {rho:+.3f}, p = {p:.4f}). The classification is "
              f"therefore not an artefact of the flow data: a split computed "
              f"without any chemical input lines up with the mechanism that "
              f"removes substances from river water.\n")
        elif p >= 0.05:
            A(f"> **No relationship detected** (rho = {rho:+.3f}, "
              f"p = {p:.4f}). The mass-balance split is not corroborated by "
              f"hydrophobicity here. With {logp[1]} substances the test is "
              f"weak, so this is not evidence of absence -- but it is not "
              f"support either, and the classification should be read as "
              f"hydrological bookkeeping rather than as evidence about "
              f"mechanism. We report it because a test run only when it "
              f"succeeds is not a test.\n")
        else:
            A(f"> **The relationship runs the wrong way** "
              f"(rho = {rho:+.3f}, p = {p:.4f}), which counts against reading "
              f"the split as a removal signal. Reported as found.\n")

    if rows:
        A("## Most and least attenuated substances\n")
        ranked = sorted(rows, key=lambda r: (-float(r["attenuation_fraction"]),
                                             -int(r["n_comparisons"])))
        A("| analyte | comparisons | requiring attenuation | log P | "
          "solubility (ppm) |")
        A("|---|---|---|---|---|")
        for r in ranked[:8]:
            A(f"| {r['analyte']} | {r['n_comparisons']} | "
              f"{100*float(r['attenuation_fraction']):.0f}% | "
              f"{r['logp'] or '—'} | {r['solubility_ppm'] or '—'} |")
        if len(ranked) > 12:
            A("| … | | | | |")
        for r in ranked[-4:]:
            A(f"| {r['analyte']} | {r['n_comparisons']} | "
              f"{100*float(r['attenuation_fraction']):.0f}% | "
              f"{r['logp'] or '—'} | {r['solubility_ppm'] or '—'} |")
        A("")

    A("## Limits\n")
    A("- log P and solubility are the vendor's reported values, not measured "
      "in this study.")
    A("- Sorption depends on suspended solids and organic carbon, neither of "
      "which was measured, so hydrophobicity is a proxy for the mechanism, "
      "not a model of it.")
    A(f"- {len(rows)} substances is a small sample; only a strong effect "
      f"would be detectable.\n")

    text = "\n".join(L)
    (EVAL / "attenuation_chemistry.md").write_text(text, encoding="utf-8")
    print(text)
    print(f"\nwrote: {EVAL/'attenuation_chemistry.md'}"
          + (f", {PROC/'attenuation_chemistry.csv'}" if rows else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

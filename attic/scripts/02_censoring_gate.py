#!/usr/bin/env python3
"""
Gate check for the paper's central claims.

Answers three questions, all of which must be settled before writing:

  Q1  How much of the campaign data is censored-but-recorded-as-zero?
      (i.e. how large is the "fabricated zero" problem, quantitatively)

  Q2  Can the censoring be RECOVERED? For how many measured analytes do we
      hold an LOD and an LOQ?

  Q3  Do analytes exist for which LOQ > EQS -- i.e. the analytical method
      cannot demonstrate compliance even in principle?
      If the answer is "none", the IndeterminateCompliance sub-claim is
      dropped from the paper. This script is the arbiter.

Inputs  (read-only):  Data/CKS_FEnCY-2.xlsx  (sheet 'LOD&LOQ')
                      Data/CKS_FEnCY.xlsx    (sheet 'CKS')
                      derived/processed/measurements.csv   (from script 01)
Outputs:              derived/processed/analytes.csv
                      eval/gate_check.md

Usage:  python scripts/02_censoring_gate.py
"""

from __future__ import annotations

import csv
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "Data"
PROC = ROOT / "derived" / "processed"
EVAL = ROOT / "eval"


def clean(s) -> str:
    """Normalise a chemical name for joining across files.

    Handles the non-breaking spaces, trailing blanks and case differences that
    the source spreadsheets are full of. Deliberately conservative: it does not
    strip punctuation, because '2,4-' vs '24-' would collide.
    """
    if s is None:
        return ""
    text = unicodedata.normalize("NFKC", str(s))
    text = text.replace("\xa0", " ").replace("‑", "-").replace("‐", "-")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def key(s) -> str:
    """Join key: cleaned, lowercased, spaces removed."""
    return clean(s).lower().replace(" ", "")


def num(s):
    """Parse a numeric cell that may carry stray unicode spaces."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    text = clean(s).replace(",", ".")
    text = re.sub(r"[^0-9eE.+-]", "", text)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_lod_loq() -> dict:
    """sheet 'LOD&LOQ': Chemical Name | Slope | LOD | LOQ | R2 | 4x EQS cols."""
    path = DATA / "CKS_FEnCY-2.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["LOD&LOQ"]
    out = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = clean(row[0])
        if not name:
            continue
        out[key(name)] = {
            "name": name,
            "slope": num(row[1]),
            "lod": num(row[2]),
            "loq": num(row[3]),
            "r2": num(row[4]),
            "eqs_aa_river": num(row[5]) if len(row) > 5 else None,
            "eqs_max_river": num(row[6]) if len(row) > 6 else None,
        }
    wb.close()
    return out


def load_eqs() -> dict:
    """sheet 'CKS': Kirletici Adi | YO-CKS (annual avg) | MAK-CKS (max allowable)."""
    path = DATA / "CKS_FEnCY.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["CKS"]
    out = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = clean(row[0])
        if not name:
            continue
        out[key(name)] = {"name": name,
                          "eqs_aa": num(row[1]),
                          "eqs_max": num(row[2])}
    wb.close()
    return out


def load_measurements():
    path = PROC / "measurements.csv"
    if not path.exists():
        sys.exit("run scripts/01_extract_campaigns.py first")
    with path.open(encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def main() -> int:
    EVAL.mkdir(parents=True, exist_ok=True)
    lodloq = load_lod_loq()
    eqs = load_eqs()
    rows = load_measurements()

    # ---- per-parameter censoring statistics -------------------------------
    stats = defaultdict(lambda: {"n": 0, "zero": 0, "pos": 0, "neg": 0,
                                 "group": "", "unit": "", "min_pos": None})
    for r in rows:
        p = r["parameter"]
        s = stats[p]
        s["group"] = r["group"]
        s["unit"] = r["unit"]
        if r["value_num"] == "":
            continue
        v = float(r["value_num"])
        s["n"] += 1
        if v == 0:
            s["zero"] += 1
        elif v < 0:
            s["neg"] += 1
        else:
            s["pos"] += 1
            if s["min_pos"] is None or v < s["min_pos"]:
                s["min_pos"] = v

    # ---- join ------------------------------------------------------------
    analytes = []
    for p in sorted(stats):
        s = stats[p]
        k = key(p)
        ll = lodloq.get(k)
        eq = eqs.get(k)
        eqs_aa = (eq or {}).get("eqs_aa")
        eqs_max = (eq or {}).get("eqs_max")
        if eqs_aa is None and ll:
            eqs_aa = ll.get("eqs_aa_river")
        if eqs_max is None and ll:
            eqs_max = ll.get("eqs_max_river")
        loq = (ll or {}).get("loq")
        analytes.append({
            "parameter": p, "group": s["group"], "unit": s["unit"],
            "n": s["n"], "n_zero": s["zero"], "n_pos": s["pos"], "n_neg": s["neg"],
            "pct_zero": round(100 * s["zero"] / s["n"], 2) if s["n"] else "",
            "min_positive": s["min_pos"] if s["min_pos"] is not None else "",
            "lod": (ll or {}).get("lod", ""), "loq": loq if loq is not None else "",
            "r2": (ll or {}).get("r2", ""),
            "eqs_aa": eqs_aa if eqs_aa is not None else "",
            "eqs_max": eqs_max if eqs_max is not None else "",
            "has_lodloq": bool(ll and ll.get("lod") is not None),
            "has_eqs": eqs_aa is not None or eqs_max is not None,
            "loq_gt_eqs_aa": (loq is not None and eqs_aa is not None and loq > eqs_aa),
            "loq_gt_eqs_max": (loq is not None and eqs_max is not None and loq > eqs_max),
        })

    with (PROC / "analytes.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(analytes[0].keys()))
        w.writeheader()
        w.writerows(analytes)

    # ---- report ----------------------------------------------------------
    micro = [a for a in analytes if a["group"] == "micropollutant"]
    measured_chem = [a for a in analytes
                     if a["group"] in ("micropollutant", "metal", "conventional")]

    tot_n = sum(a["n"] for a in measured_chem)
    tot_zero = sum(a["n_zero"] for a in measured_chem)
    micro_n = sum(a["n"] for a in micro)
    micro_zero = sum(a["n_zero"] for a in micro)

    with_lodloq = [a for a in measured_chem if a["has_lodloq"]]
    with_eqs = [a for a in measured_chem if a["has_eqs"]]
    recoverable = sum(a["n_zero"] for a in with_lodloq)

    conflict_aa = [a for a in analytes if a["loq_gt_eqs_aa"]]
    conflict_max = [a for a in analytes if a["loq_gt_eqs_max"]]

    # analytes whose smallest reported positive value is below their own LOQ
    below_loq = [a for a in with_lodloq
                 if a["min_positive"] != "" and a["loq"] != ""
                 and a["min_positive"] < a["loq"]]

    L = []
    A = L.append
    A("# Gate check\n")
    A("Generated by `scripts/02_censoring_gate.py`. "
      "Decides whether the paper's central claims survive contact with the data.\n")

    A("## Q1 — How large is the 'zero-substituted non-detect' problem?\n")
    A(f"- Chemical measurements (conventional + metal + micropollutant): **{tot_n:,}**")
    A("")
    # Report the breakdown this script already computes. Leaving it out meant
    # the per-group shares existed only inside the code, so nothing downstream
    # could be checked against them.
    A("| determinand group | measurements | stored as 0.0 | share |")
    A("|---|---|---|---|")
    for _g, _lab in (("conventional", "Conventional"), ("metal", "Metals"),
                     ("micropollutant", "Micropollutants")):
        _rows = [a for a in analytes if a["group"] == _g]
        if not _rows:
            continue
        _n = sum(a["n"] for a in _rows)
        _z = sum(a["n_zero"] for a in _rows)
        A(f"| {_lab} | {_n:,} | {_z:,} | {100*_z/_n:.1f}% |")
    A("")
    A(f"- Recorded as exactly `0.0`: **{tot_zero:,} ({100*tot_zero/tot_n:.1f}%)**")
    A(f"- Micropollutants alone: {micro_zero:,} / {micro_n:,} "
      f"(**{100*micro_zero/micro_n:.1f}%**)")
    A("")
    A("A reported `0.0` is not a measurement of zero concentration; it is a "
      "non-detect with its censoring information erased. Every load, mean and "
      "sum computed from this table is biased low by an unknown amount "
      "(Helsel 2006).\n")

    A("## Q2 — Can the censoring be recovered?\n")
    A(f"- Analytes with LOD **and** LOQ available: "
      f"**{len(with_lodloq)}** of {len(measured_chem)} measured chemical analytes")
    A(f"- Zero-records these cover: **{recoverable:,}** "
      f"({100*recoverable/tot_zero:.1f}% of all zeros)")
    A(f"- Analytes with an EQS: **{len(with_eqs)}**")
    A("")
    if len(with_lodloq) < len(measured_chem):
        A(f"> {len(measured_chem) - len(with_lodloq)} measured analytes have no "
          f"LOD/LOQ in `CKS_FEnCY-2.xlsx`. For these, censoring cannot be "
          f"reconstructed and the reach verdict must be `UndecidableReach` by "
          f"construction — which is itself a reportable result.\n")

    A("## Q3 — GATE: do analytes exist where LOQ > EQS?\n")
    if conflict_aa or conflict_max:
        A(f"**YES — the sub-claim survives.**\n")
        A(f"- LOQ > annual-average EQS: **{len(conflict_aa)}** analytes")
        A(f"- LOQ > maximum-allowable EQS: **{len(conflict_max)}** analytes\n")
        A("| Analyte | LOQ (µg/L) | EQS-AA | EQS-MAX | LOQ>AA | LOQ>MAX |")
        A("|---|---|---|---|---|---|")
        seen = {a["parameter"] for a in conflict_aa} | {a["parameter"] for a in conflict_max}
        for a in sorted(analytes, key=lambda x: -(x["loq"] or 0) if x["loq"] != "" else 0):
            if a["parameter"] not in seen:
                continue
            A(f"| {a['parameter']} | {a['loq']:.4g} | "
              f"{a['eqs_aa'] if a['eqs_aa']!='' else '—'} | "
              f"{a['eqs_max'] if a['eqs_max']!='' else '—'} | "
              f"{'YES' if a['loq_gt_eqs_aa'] else ''} | "
              f"{'YES' if a['loq_gt_eqs_max'] else ''} |")
        A("")
        A("For these substances a 'compliant' result is an artefact of the "
          "analytical method, not evidence of compliance. This is precisely the "
          "`IndeterminateCompliance` case.\n")
    else:
        A("**NO — no analyte has LOQ above its EQS.**\n")
        A("Drop the `IndeterminateCompliance` sub-claim based on LOQ>EQS. "
          "The censoring contribution stands on Q1/Q2 alone.\n")

    A("## Q4 — Internal consistency of the reported values\n")
    A(f"- Analytes whose smallest reported positive value lies **below their own "
      f"LOQ**: **{len(below_loq)}**")
    if below_loq:
        A("  These are values reported as if quantified while being below the "
          "quantification limit — they should be `EstimatedObservation`, not "
          "point values. Examples:")
        for a in sorted(below_loq, key=lambda x: x["min_positive"])[:10]:
            A(f"    - {a['parameter']}: min positive {a['min_positive']:.4g} "
              f"< LOQ {a['loq']:.4g}")
    neg = [a for a in analytes if a["n_neg"]]
    if neg:
        A(f"- Analytes with **negative** concentrations (physically impossible): "
          f"{', '.join(a['parameter'] for a in neg)}")
    A("")

    A("## Name-join quality\n")
    unmatched = [a["parameter"] for a in measured_chem if not a["has_lodloq"]]
    A(f"- Measured chemical analytes not matched to LOD/LOQ: {len(unmatched)}")
    A(f"- LOD/LOQ entries in file: {len(lodloq)} · EQS entries: {len(eqs)}")
    A("- Unmatched sample (first 25):")
    for u in unmatched[:25]:
        A(f"  - {u}")
    A("")
    A("> Join is exact-after-normalisation (NFKC, non-breaking spaces removed, "
      "lowercased, spaces stripped). Fuzzy matching is deliberately **not** used: "
      "a wrong chemical match silently corrupts a load budget. Unmatched names "
      "are resolved manually into a mapping table in WP4.\n")

    text = "\n".join(L)
    (EVAL / "gate_check.md").write_text(text, encoding="utf-8")
    print(text)
    print(f"\nwrote: {EVAL/'gate_check.md'}, {PROC/'analytes.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

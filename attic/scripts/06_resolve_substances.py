#!/usr/bin/env python3
"""
Resolve measured analytes to external substance identifiers and USE CATEGORIES.

WHY THIS MATTERS FOR THE ARGUMENT
---------------------------------
Source attribution works by asking which substances first appear in a reach and
what those substances are used for. If the analyst writes the
substance-to-use-class mapping themselves, and then validates the result against
land use, the validation is CIRCULAR: the answer was built into the mapping.

So the mapping is taken from external, citable, versioned sources only:

  * PubChem CID          - identity anchor (PUG REST)
  * ChEBI                - chemical ontology alignment, for skos:exactMatch
  * EPA CPDat            - "Function Use Categories" (Herbicide, Biocide, ...)
                           and Product Use Categories: the USE evidence
  * MeSH                 - pharmacological classes for pharmaceuticals
  * ECOSAR class         - structural class, already held locally

Nothing here is authored by us. Every use category carries the source it came
from, so a reviewer can check any single assignment.

REPRODUCIBILITY
---------------
Raw API responses are cached under derived/interim/pubchem_cache/. A second run
uses the cache and needs no network, so the published results are reproducible
after PubChem changes. Delete the cache to refresh.

PubChem asks for <= 5 requests/second; this script stays well under that.

Inputs  : derived/processed/analytes.csv, Data/Toxicity_Ecosar_*.xlsx,
          Data/Properties_report*.xlsx
Outputs : derived/processed/substances.csv
          derived/interim/pubchem_cache/*.json
          eval/substance_resolution.md

Usage:  python scripts/06_resolve_substances.py [--limit N] [--offline]
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "Data"
PROC = ROOT / "derived" / "processed"
CACHE = ROOT / "derived" / "interim" / "pubchem_cache"
EVAL = ROOT / "eval"

PUG = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
DELAY = 0.25          # seconds between requests
TIMEOUT = 40
UA = "censo-ontology-research/1.0 (academic use; contact via repository)"

# Sources we accept a USE category from. Structural hierarchies (ChEBI, ChEMBL)
# describe what a molecule IS; these describe what it is USED FOR, which is what
# source attribution needs.
USE_SOURCES = {
    "EPA Chemical and Products Database (CPDat)": "cpdat",
    "Medical Subject Headings (MeSH)": "mesh",
    "Consumer Product Information Database (CPID)": "cpid",
}

# Metal symbols as measured -> full name, so they resolve in PubChem.
METAL_NAMES = {
    "Ag": "Silver", "Al": "Aluminum", "As": "Arsenic", "B": "Boron",
    "Ba": "Barium", "Be": "Beryllium", "Cd": "Cadmium", "Co": "Cobalt",
    "Cr": "Chromium", "Cu": "Copper", "Hg": "Mercury", "Ni": "Nickel",
    "Pb": "Lead", "Sb": "Antimony", "Sn": "Tin", "Ti": "Titanium",
    "V": "Vanadium", "Zn": "Zinc",
}


def key(s) -> str:
    s = unicodedata.normalize("NFKC", str(s)).replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip().lower().replace(" ", "")


COMPACT = CACHE / "classification_compact.json"


def load_compact() -> dict:
    if COMPACT.exists():
        try:
            return json.loads(COMPACT.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {}
    return {}


def save_compact(d: dict) -> None:
    CACHE.mkdir(parents=True, exist_ok=True)
    COMPACT.write_text(json.dumps(d, ensure_ascii=False, indent=0, sort_keys=True),
                       encoding="utf-8")


def fetch_classification(cid: str, compact: dict, offline: bool):
    """Return {'chebi': str, 'uses': {source: [terms]}} for a CID.

    The raw response is parsed and discarded immediately; only the extracted
    fields are kept.
    """
    if cid in compact:
        return compact[cid]
    if offline:
        return None
    url = (f"{PUG}/compound/cid/{cid}/classification/JSON"
           f"?classification_type=simple")
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
            raw = r.read().decode("utf-8", errors="replace")
    except Exception:
        return None
    finally:
        time.sleep(DELAY)
    try:
        d = json.loads(raw)
    except json.JSONDecodeError:
        return None

    rec = {"chebi": "", "uses": {}}
    for h in ((d or {}).get("Hierarchies", {}) or {}).get("Hierarchy", []) or []:
        src = str(h.get("SourceName", ""))
        if src == "ChEBI" and not rec["chebi"]:
            rec["chebi"] = str(h.get("SourceID", ""))
        if src in USE_SOURCES:
            tag = USE_SOURCES[src]
            terms = sorted(set(collect_terms(h, USE_BRANCHES.get(tag, set()))))
            if terms:
                rec["uses"][tag] = terms
    compact[cid] = rec
    return rec


def fetch(url: str, cache_name: str, offline: bool):
    """GET with an on-disk cache. Returns parsed JSON or None."""
    CACHE.mkdir(parents=True, exist_ok=True)
    path = CACHE / f"{cache_name}.json"
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8", errors="replace"))
        except json.JSONDecodeError:
            return None
    if offline:
        return None
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
            raw = r.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        if e.code == 404:
            path.write_text("null", encoding="utf-8")
            return None
        return None
    except Exception:
        return None
    finally:
        time.sleep(DELAY)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return None
    path.write_text(raw, encoding="utf-8")
    return data


def node_name(node) -> str:
    """Extract a node label.

    PubChem is inconsistent here: Name is sometimes a plain string, sometimes
    {"String": ...}, and sometimes {"StringWithMarkup": ...} where the inner
    value is EITHER a dict OR a list of dicts. Handling only the list form
    silently yields empty labels, which looks like "this substance has no use
    category" rather than like a parsing bug.
    """
    v = node.get("Information", {}).get("Name", "")
    if isinstance(v, str):
        return v
    if not isinstance(v, dict):
        return ""
    if "String" in v:
        return str(v["String"])
    sw = v.get("StringWithMarkup")
    if isinstance(sw, dict):
        return str(sw.get("String", ""))
    if isinstance(sw, list) and sw and isinstance(sw[0], dict):
        return str(sw[0].get("String", ""))
    return ""


# Only these branches carry USE information. The rest of each hierarchy is
# navigational scaffolding ("List Presence Keywords", "Modifiers") or regulatory
# list membership ("OEHHA Proposition 65"), which says nothing about what a
# substance is used for.
USE_BRANCHES = {
    "cpdat": {"Function Use Categories", "Product Use Categories (PUCs)"},
    "mesh": {"Chemical Actions and Uses"},
    "cpid": {"Product Category"},
}


def collect_terms(hierarchy, branch_roots):
    """Leaf labels beneath the named branch roots.

    PubChem returns each hierarchy as a FLAT list of nodes carrying NodeID and
    ParentID, not as nested objects. The tree is rebuilt here so that a branch
    can be isolated; walking the flat list instead returns every label in the
    hierarchy, including scaffolding, which is how a parser bug masquerades as
    a rich result.
    """
    nodes = hierarchy.get("Node") or []
    if not isinstance(nodes, list):
        return []

    by_id, children, label = {}, defaultdict(list), {}
    for n in nodes:
        if not isinstance(n, dict):
            continue
        nid = n.get("NodeID")
        by_id[nid] = n
        label[nid] = node_name(n)
        # ParentID may be a single id or a list: these classifications are DAGs,
        # not trees, because a substance use can sit under several headings.
        pid = n.get("ParentID")
        parents = pid if isinstance(pid, list) else ([] if pid is None else [pid])
        for pp in parents:
            if isinstance(pp, (str, int)):
                children[pp].append(nid)

    roots = [nid for nid, lab in label.items() if lab in branch_roots]
    if not roots:
        return []

    out, seen = [], set()

    def walk(nid, depth=0):
        if nid in seen or depth > 12:
            return
        seen.add(nid)
        kids = children.get(nid, [])
        if not kids:
            lab = label.get(nid, "")
            # the branch root itself is not a use category
            if lab and lab not in branch_roots:
                out.append(lab)
        for k in kids:
            walk(k, depth + 1)

    for r in roots:
        walk(r)
    return out


def load_local():
    """CAS, SMILES and ECOSAR class from the project spreadsheets."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required")

    ecosar = {}
    wb = openpyxl.load_workbook(DATA / "Toxicity_Ecosar_FINAL-2018_03_12.xlsx",
                                data_only=True, read_only=True)
    for r in list(wb["Batch Output"].iter_rows(values_only=True))[1:]:
        if r[1] is None:
            continue
        ecosar[key(r[1])] = {
            "name": str(r[1]).strip(),
            "cas": str(r[0]).strip() if r[0] else "",
            "smiles": str(r[2]).strip() if r[2] else "",
            "ecosar_class": str(r[3]).strip() if r[3] else "",
        }
    wb.close()

    props = {}
    wb = openpyxl.load_workbook(DATA / "Properties_report order_2018-07-04.xlsx",
                                data_only=True, read_only=True)
    for sheet, ni, si in [("PROPERTIES", 2, 3), ("Metals", 1, 2)]:
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
            if r[ni] is None:
                continue
            props[key(r[ni])] = {"name": str(r[ni]).strip(),
                                 "smiles": str(r[si]).strip() if r[si] else ""}
    wb.close()
    return ecosar, props


def normalise_cas(cas: str) -> str:
    """CAS numbers appear both with and without hyphens in the source files."""
    c = re.sub(r"[^0-9-]", "", cas or "")
    if not c:
        return ""
    if "-" in c:
        return c
    if len(c) >= 5:                      # 50000 -> 50-00-0
        return f"{c[:-3]}-{c[-3:-1]}-{c[-1]}"
    return c


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, help="resolve only the first N analytes")
    ap.add_argument("--offline", action="store_true",
                    help="use only the cache; make no network requests")
    args = ap.parse_args()
    EVAL.mkdir(parents=True, exist_ok=True)

    compact = load_compact()
    ecosar, props = load_local()
    analytes = [a for a in csv.DictReader((PROC / "analytes.csv").open(encoding="utf-8"))
                if a["group"] in ("micropollutant", "metal")]
    if args.limit:
        analytes = analytes[:args.limit]

    rows = []
    stats = Counter()
    use_counter = Counter()

    for a in analytes:
        name = a["parameter"]
        k = key(name)
        e = ecosar.get(k, {})
        p = props.get(k, {})

        lookup = METAL_NAMES.get(name, name)
        cas = normalise_cas(e.get("cas", ""))
        smiles = e.get("smiles") or p.get("smiles") or ""

        # CID: try CAS first (unambiguous), then the name.
        cid = ""
        for query, tag in ((cas, "cas"), (lookup, "name")):
            if not query:
                continue
            safe = urllib.parse.quote(query, safe="")
            d = fetch(f"{PUG}/compound/name/{safe}/cids/JSON",
                      f"cid_{tag}_{re.sub(r'[^A-Za-z0-9._-]', '_', query)[:80]}",
                      args.offline)
            ids = (d or {}).get("IdentifierList", {}).get("CID", []) if d else []
            if ids:
                cid = str(ids[0])
                stats[f"cid_via_{tag}"] += 1
                break
        if not cid:
            stats["cid_unresolved"] += 1

        chebi, uses = "", defaultdict(set)
        if cid:
            rec = fetch_classification(cid, compact, args.offline)
            if rec:
                chebi = rec.get("chebi", "")
                for tag, terms in (rec.get("uses") or {}).items():
                    uses[tag].update(terms)
            if chebi:
                stats["chebi_aligned"] += 1
            if uses:
                stats["use_categories_found"] += 1

        flat = sorted({f"{s}:{t}" for s, ts in uses.items() for t in ts})
        for f in flat:
            use_counter[f] += 1

        rows.append({
            "analyte": name, "group": a["group"],
            "cas": cas, "pubchem_cid": cid, "chebi_id": chebi,
            "smiles": smiles, "ecosar_class": e.get("ecosar_class", ""),
            "n_use_categories": len(flat),
            "use_categories": " | ".join(flat),
        })

    save_compact(compact)

    with (PROC / "substances.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    n = len(rows)
    L = []
    A = L.append
    A("# Substance resolution to external identifiers and use categories\n")
    A("Generated by `scripts/06_resolve_substances.py`. Every use category comes "
      "from an external, citable source; none is authored here. This is what "
      "keeps the land-use validation non-circular.\n")
    A("## Coverage\n")
    A("| step | n | share |")
    A("|---|---|---|")
    A(f"| analytes attempted | {n} | 100% |")
    A(f"| CAS available locally | {sum(1 for r in rows if r['cas'])} | "
      f"{100*sum(1 for r in rows if r['cas'])/n:.1f}% |")
    A(f"| PubChem CID via CAS | {stats['cid_via_cas']} | {100*stats['cid_via_cas']/n:.1f}% |")
    A(f"| PubChem CID via name | {stats['cid_via_name']} | {100*stats['cid_via_name']/n:.1f}% |")
    A(f"| **CID unresolved** | **{stats['cid_unresolved']}** | {100*stats['cid_unresolved']/n:.1f}% |")
    A(f"| ChEBI aligned | {stats['chebi_aligned']} | {100*stats['chebi_aligned']/n:.1f}% |")
    A(f"| **with >=1 use category** | **{stats['use_categories_found']}** | "
      f"{100*stats['use_categories_found']/n:.1f}% |")
    A("")
    A("## Most frequent use categories\n")
    A("| category | source | analytes |")
    A("|---|---|---|")
    for term, c in use_counter.most_common(30):
        src, _, label = term.partition(":")
        A(f"| {label} | {src} | {c} |")
    A("")
    A("Analytes with no use category are reported as such and excluded from "
      "source-type inference rather than assigned a default class.\n")
    A("## Provenance\n")
    A("- PubChem PUG REST, `classification_type=simple`\n"
      "- Use categories accepted only from: "
      + ", ".join(sorted(USE_SOURCES)) + "\n"
      "- ChEBI used for ontology alignment (`skos:exactMatch`), not for use\n"
      "- Raw responses cached under `derived/interim/pubchem_cache/` so the "
      "resolution is reproducible without network access\n")

    text = "\n".join(L)
    (EVAL / "substance_resolution.md").write_text(text, encoding="utf-8")
    print(text)
    print(f"\nwrote: {EVAL/'substance_resolution.md'}, {PROC/'substances.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
